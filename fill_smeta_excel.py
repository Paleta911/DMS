from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import pythoncom
import win32com.client as win32
from openpyxl import load_workbook


BASE_DIR = Path(r"C:\Users\alexi\Downloads\originales")
SOURCE_PATH = BASE_DIR / "Avance Plan Accion SMETA-CEP NESTLE.xlsx"
DEST_TEMPLATE_PATH = BASE_DIR / "CEP-FOR-RC-17 REV.08  SEDEX SMETA.xlsx"
OUTPUT_PATH = BASE_DIR / "CEP-FOR-RC-17 REV.08 SEDEX SMETA_Llenado.xlsx"

SOURCE_SHEET = "PlanAccion"
DEST_SHEET = "CEP-FOR-RC-17 Rev. 08"

START_ROW = 16
ROWS_PER_RECORD = 4
EXPECTED_RECORDS = 79
RESPONSIBLE_BY_PILLAR = {
    "Seguridad e Higiene": "Ing. Diego Rodr\u00edguez Garc\u00eda",
    "Laboral": "Lic. Jos\u00e9 De Jes\u00fas Acosta",
    "Ambiental": "Q.F.B. Jos\u00e9 Mart\u00ednez Ram\u00edrez",
    "\u00c9tica": "Lic. Celia Cervantes Reyna",
    "GAP": "Lic. Jos\u00e9 De Jes\u00fas Acosta",
}

CORRECTION_LABEL_RE = re.compile(
    r"corre(?:cc|c)?i(?:o|\u00f3)n(?:es)?(?:\s+a\s+realizar)?\s*[:.\-]+",
    re.IGNORECASE,
)
ACTION_LABEL_RE = re.compile(
    r"acci(?:o|\u00f3)n(?:\s+(?:correctiva|(?:u|\u00fa)nica))?\s*[:.\-]+",
    re.IGNORECASE,
)


@dataclass
class Record:
    number: int
    pillar: str
    finding: str
    root_cause: str
    correction: str
    corrective_action: str
    responsible: str


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    return text.strip()


def split_action_text(value: object) -> tuple[str, str]:
    text = normalize_text(value)
    if not text:
        return "", ""

    correction_match = CORRECTION_LABEL_RE.search(text)
    action_match = ACTION_LABEL_RE.search(text)

    correction = ""
    corrective_action = ""

    if correction_match and action_match:
        if correction_match.start() < action_match.start():
            correction = text[correction_match.end() : action_match.start()]
            corrective_action = text[action_match.end() :]
        else:
            corrective_action = text[action_match.end() : correction_match.start()]
            correction = text[correction_match.end() :]
    elif correction_match:
        correction = text[correction_match.end() :]
    elif action_match:
        corrective_action = text[action_match.end() :]
    else:
        corrective_action = text

    return normalize_text(correction), normalize_text(corrective_action)


def load_source_records() -> list[Record]:
    workbook = load_workbook(SOURCE_PATH, data_only=True, read_only=True)
    try:
        sheet = workbook[SOURCE_SHEET]
        records: list[Record] = []
        for row in range(6, sheet.max_row + 1):
            number = sheet.cell(row, 2).value
            pillar = sheet.cell(row, 3).value
            finding = sheet.cell(row, 4).value
            if number is None and pillar is None and finding is None:
                continue

            correction, corrective_action = split_action_text(sheet.cell(row, 10).value)
            records.append(
                Record(
                    number=int(number),
                    pillar=normalize_text(pillar),
                    finding=normalize_text(finding),
                    root_cause=normalize_text(sheet.cell(row, 9).value),
                    correction=correction,
                    corrective_action=corrective_action,
                    responsible=RESPONSIBLE_BY_PILLAR.get(normalize_text(pillar), ""),
                )
            )
    finally:
        workbook.close()

    numbers = sorted(record.number for record in records)
    expected_numbers = list(range(1, EXPECTED_RECORDS + 1))
    if numbers != expected_numbers:
        raise RuntimeError(
            f"Source records do not match expected numbering: {numbers[:5]} ... {numbers[-5:]}"
        )

    return records


def copy_template() -> None:
    if OUTPUT_PATH.exists():
        OUTPUT_PATH.unlink()
    shutil.copy2(DEST_TEMPLATE_PATH, OUTPUT_PATH)


def set_cell(sheet, address: str, value: str) -> None:
    sheet.Range(address).Value = value if value else None


def validate_destination_layout(sheet) -> None:
    for number in range(1, EXPECTED_RECORDS + 1):
        row = START_ROW + (number - 1) * ROWS_PER_RECORD
        current = sheet.Range(f"A{row}").Value
        if current != number:
            raise RuntimeError(
                f"Destination block mismatch at row {row}: expected {number}, found {current!r}"
            )


class HeightMeasurer:
    def __init__(self, excel):
        self.workbook = excel.Workbooks.Add()
        self.sheet = self.workbook.Worksheets(1)
        self.base_column_width = self.sheet.Columns(1).ColumnWidth
        self.base_width_points = self.sheet.Columns(1).Width
        self.standard_row_height = float(self.sheet.StandardHeight)

    def close(self) -> None:
        if self.workbook is not None:
            self.workbook.Close(SaveChanges=False)
            self.workbook = None
            self.sheet = None

    def _set_width_points(self, target_points: float) -> None:
        column = self.sheet.Columns(1)
        column.ColumnWidth = self.base_column_width * target_points / self.base_width_points
        for _ in range(12):
            current_width = float(column.Width)
            if abs(current_width - target_points) < 0.2:
                break
            column.ColumnWidth = column.ColumnWidth * (target_points / current_width)

    def measure_height(self, template_range, text: str) -> float:
        if not text:
            return 0.0

        self.sheet.Cells.Clear()
        self.sheet.Rows(1).RowHeight = self.standard_row_height
        self._set_width_points(float(template_range.MergeArea.Width))

        cell = self.sheet.Cells(1, 1)
        cell.Value = text
        cell.WrapText = True
        cell.Font.Name = template_range.Font.Name
        cell.Font.Size = template_range.Font.Size
        cell.Font.Bold = template_range.Font.Bold
        cell.Font.Italic = template_range.Font.Italic
        cell.Font.Underline = template_range.Font.Underline
        cell.HorizontalAlignment = template_range.HorizontalAlignment
        cell.VerticalAlignment = template_range.VerticalAlignment
        cell.Orientation = template_range.Orientation
        cell.IndentLevel = template_range.IndentLevel
        cell.ShrinkToFit = False

        self.sheet.Rows(1).AutoFit()
        height = float(self.sheet.Rows(1).RowHeight)

        if bool(template_range.MergeCells):
            height += max(8.0, float(template_range.Font.Size) * 1.3)

        return height


def adjust_row_heights(sheet, measurer: HeightMeasurer, record: Record) -> None:
    row_with_finding = START_ROW + (record.number - 1) * ROWS_PER_RECORD + 1
    row_with_cause = row_with_finding + 2

    finding_height = measurer.measure_height(sheet.Range(f"C{row_with_finding}"), record.finding)
    correction_height = measurer.measure_height(
        sheet.Range(f"I{row_with_finding}"), record.correction
    )
    target_finding_row_height = max(
        float(sheet.Rows(row_with_finding).RowHeight),
        finding_height,
        correction_height,
    )
    sheet.Rows(row_with_finding).RowHeight = target_finding_row_height

    cause_height = measurer.measure_height(sheet.Range(f"C{row_with_cause}"), record.root_cause)
    action_height = measurer.measure_height(
        sheet.Range(f"I{row_with_cause}"), record.corrective_action
    )
    target_cause_row_height = max(
        float(sheet.Rows(row_with_cause).RowHeight),
        cause_height,
        action_height,
    )
    sheet.Rows(row_with_cause).RowHeight = target_cause_row_height


def fill_destination(records: list[Record]) -> None:
    pythoncom.CoInitialize()
    excel = None
    workbook = None
    measurer = None

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False

        workbook = excel.Workbooks.Open(str(OUTPUT_PATH), UpdateLinks=0, ReadOnly=False)
        sheet = workbook.Worksheets(DEST_SHEET)
        measurer = HeightMeasurer(excel)
        validate_destination_layout(sheet)

        for record in records:
            base_row = START_ROW + (record.number - 1) * ROWS_PER_RECORD
            set_cell(sheet, f"A{base_row}", str(record.number))
            set_cell(sheet, f"B{base_row}", record.pillar)
            set_cell(sheet, f"C{base_row + 1}", record.finding)
            set_cell(sheet, f"C{base_row + 3}", record.root_cause)
            set_cell(sheet, f"I{base_row + 1}", record.correction)
            set_cell(sheet, f"I{base_row + 3}", record.corrective_action)
            set_cell(sheet, f"K{base_row}", record.responsible)
            adjust_row_heights(sheet, measurer, record)

        workbook.Save()
        workbook.Close(SaveChanges=True)
        workbook = None
        measurer.close()
        measurer = None
        excel.Quit()
        excel = None
    finally:
        if measurer is not None:
            measurer.close()
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def verify_output() -> None:
    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(OUTPUT_PATH), UpdateLinks=0, ReadOnly=True)
        sheet = workbook.Worksheets(DEST_SHEET)
        if sheet.Range("A16").Value != 1 or sheet.Range("A328").Value != 79:
            raise RuntimeError("Verification failed: record anchors do not match expected values.")
        for number in range(1, EXPECTED_RECORDS + 1):
            base_row = START_ROW + (number - 1) * ROWS_PER_RECORD
            if str(sheet.Range(f"I{base_row + 3}").Value or "").startswith("[REVISAR]"):
                raise RuntimeError(f"Verification failed: unexpected [REVISAR] found in record {number}.")
        workbook.Close(SaveChanges=False)
        workbook = None
        excel.Quit()
        excel = None
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def main() -> None:
    records = load_source_records()
    unknown_pillars = sorted({record.pillar for record in records if not record.responsible})
    copy_template()
    fill_destination(records)
    verify_output()

    print(f"Created: {OUTPUT_PATH}")
    print(f"Records processed: {len(records)}")
    print("Marked with [REVISAR]: 0")
    print(f"Unknown pillars: {unknown_pillars}")


if __name__ == "__main__":
    main()
